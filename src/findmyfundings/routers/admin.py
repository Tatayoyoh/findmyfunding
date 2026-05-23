"""Admin routes for managing sources and imports."""

import io
import json
import shutil
from datetime import date
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Request, Form
from fastapi.responses import RedirectResponse, StreamingResponse

from findmyfundings.config import settings
from findmyfundings.database import get_db
from findmyfundings.services.excel_import import parse_excel, import_to_db
from findmyfundings.services.funding_repo import (
    get_all, get_by_id, get_categories, get_all_suggestions,
)
from findmyfundings.services.scraper import scrape_program
from findmyfundings.services.scheduler import monthly_scrape_job

router = APIRouter(prefix="/admin")


@router.get("/")
async def admin_index(request: Request):
    return request.app.state.templates.TemplateResponse(
        "admin/index.html", {"request": request},
    )


@router.post("/import-excel")
async def run_excel_import(request: Request):
    programs = parse_excel(settings.excel_path)
    count = await import_to_db(programs)
    return RedirectResponse(
        f"/admin/programs?message={count}+programmes+import%C3%A9s+avec+succ%C3%A8s.",
        status_code=303,
    )


@router.post("/add-source")
async def add_source(
    request: Request,
    url: str = Form(...),
    label: str = Form(""),
):
    """Add a new source URL → create a funding program with that URL.
    Runs Firecrawl extraction to prefill structured fields when possible."""
    extraction = None
    try:
        extraction = await scrape_program(program_id=0, urls=[url])
    except Exception:
        extraction = None

    source_urls_json = json.dumps(
        [{"url": url, "label": label}], ensure_ascii=False
    )

    db = await get_db()
    try:
        if extraction:
            await db.execute(
                """INSERT INTO funding_programs
                   (category, name, project_types, summary, source_urls,
                    min_amount_eur, max_amount_eur, cofinancing_pct,
                    eligibility_criteria, eligible_structures, eligible_themes,
                    fundable_axes, relevant_links, pdf_documents, tags,
                    application_type, next_deadline,
                    permanent, start_submission_date, end_submission_date)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    "Non classifié",
                    extraction.name or label or url[:80],
                    extraction.project_types,
                    extraction.summary,
                    source_urls_json,
                    extraction.min_amount_eur,
                    extraction.max_amount_eur,
                    extraction.cofinancing_pct,
                    json.dumps(extraction.eligibility_criteria, ensure_ascii=False),
                    json.dumps(extraction.eligible_structures, ensure_ascii=False),
                    json.dumps(extraction.eligible_themes, ensure_ascii=False),
                    json.dumps(extraction.fundable_axes, ensure_ascii=False),
                    json.dumps(extraction.relevant_links, ensure_ascii=False),
                    json.dumps(extraction.pdf_documents, ensure_ascii=False),
                    json.dumps(extraction.tags, ensure_ascii=False),
                    extraction.application_type,
                    str(extraction.next_deadline) if extraction.next_deadline else None,
                    extraction.permanent,
                    str(extraction.start_submission_date) if extraction.start_submission_date else None,
                    str(extraction.end_submission_date) if extraction.end_submission_date else None,
                ),
            )
            preview = (extraction.summary or extraction.name or "")[:100]
            message = f"Programme créé et analysé : {preview}"
        else:
            await db.execute(
                """INSERT INTO funding_programs
                   (category, name, source_urls)
                   VALUES (?, ?, ?)""",
                ("Non classifié", label or url[:80], source_urls_json),
            )
            message = "Programme créé mais extraction Firecrawl indisponible."
        await db.commit()
    finally:
        await db.close()

    return RedirectResponse(
        f"/admin/programs?message={quote(message)}", status_code=303,
    )


@router.post("/run-scrape")
async def run_scrape(request: Request):
    """Trigger a manual scrape of all sources."""
    await monthly_scrape_job()
    return RedirectResponse(
        "/admin/programs?message=Scraping+termin%C3%A9.", status_code=303,
    )


# --- Programs CRUD ---


def _parse_comma_list(value: str) -> list[str]:
    """Parse comma-separated string into list, stripping whitespace."""
    return [s.strip() for s in value.split(",") if s.strip()] if value else []


def _parse_source_urls(text: str, existing: list | None = None) -> str:
    """Parse newline-separated URLs into JSON list. Preserves scraping
    metadata (last_hash, last_checked_at, has_changed) when a URL exists
    in `existing`."""
    existing_by_url = {}
    if existing:
        for e in existing:
            url = e.url if hasattr(e, "url") else e.get("url")
            if url:
                existing_by_url[url] = e

    out = []
    for line in text.splitlines():
        url = line.strip()
        if not url:
            continue
        prior = existing_by_url.get(url)
        if prior is not None:
            entry = prior.model_dump(mode="json") if hasattr(prior, "model_dump") else dict(prior)
            entry["url"] = url
            entry.setdefault("label", "")
            out.append(entry)
        else:
            out.append({"url": url, "label": ""})
    return json.dumps(out, ensure_ascii=False)


@router.get("/programs")
async def programs_list(
    request: Request, message: str = "", scraping: str = "", expired: str = "",
):
    """List all funding programs for management."""
    programs = await get_all()

    truthy = ("1", "true", "on")
    scraping_only = scraping in truthy
    expired_only = expired in truthy

    today_str = date.today().isoformat()

    def is_expired(p) -> bool:
        return bool(
            p.end_submission_date
            and str(p.end_submission_date) < today_str
            and not p.permanent
        )

    if scraping_only:
        programs = [p for p in programs if p.source_urls]
    if expired_only:
        programs = [p for p in programs if is_expired(p)]

    ctx = {
        "request": request,
        "programs": programs,
        "scraping_only": scraping_only,
        "expired_only": expired_only,
    }
    if message:
        ctx["message"] = message
    return request.app.state.templates.TemplateResponse(
        "admin/programs.html", ctx,
    )


@router.get("/programs/new")
async def program_new_form(request: Request):
    """Show form to create a new program."""
    categories = await get_categories()
    suggestions = await get_all_suggestions()
    return request.app.state.templates.TemplateResponse(
        "admin/program_form.html",
        {
            "request": request,
            "program": None,
            "categories": categories,
            "suggestions": suggestions,
        },
    )


@router.post("/programs/new")
async def program_create(
    request: Request,
    name: str = Form(...),
    category: str = Form(...),
    project_types: str = Form(""),
    selection_criteria: str = Form(""),
    permanent: str = Form(""),
    start_submission_date: str = Form(""),
    end_submission_date: str = Form(""),
    pdp_axes: str = Form(""),
    comments: str = Form(""),
    min_amount_eur: Optional[int] = Form(None),
    max_amount_eur: Optional[int] = Form(None),
    cofinancing_pct: Optional[int] = Form(None),
    application_type: str = Form(""),
    next_deadline: str = Form(""),
    eligible_structures: str = Form(""),
    eligible_themes: str = Form(""),
    source_urls: str = Form(""),
):
    """Create a new funding program manually."""
    db = await get_db()
    try:
        await db.execute(
            """INSERT INTO funding_programs
               (category, name, project_types, selection_criteria,
                permanent, start_submission_date, end_submission_date,
                pdp_axes, comments, source_urls,
                min_amount_eur, max_amount_eur, cofinancing_pct,
                eligible_structures, eligible_themes,
                application_type, next_deadline)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                category,
                name,
                project_types,
                selection_criteria,
                bool(permanent),
                start_submission_date or None,
                end_submission_date or None,
                pdp_axes,
                comments,
                _parse_source_urls(source_urls),
                min_amount_eur,
                max_amount_eur,
                cofinancing_pct,
                json.dumps(_parse_comma_list(eligible_structures), ensure_ascii=False),
                json.dumps(_parse_comma_list(eligible_themes), ensure_ascii=False),
                application_type or None,
                next_deadline or None,
            ),
        )
        await db.commit()
    finally:
        await db.close()

    return RedirectResponse("/admin/programs", status_code=303)


@router.get("/programs/{program_id}/edit")
async def program_edit_form(request: Request, program_id: int):
    """Show form to edit an existing program."""
    program = await get_by_id(program_id)
    if not program:
        return RedirectResponse("/admin/programs", status_code=303)
    categories = await get_categories()
    suggestions = await get_all_suggestions()
    return request.app.state.templates.TemplateResponse(
        "admin/program_form.html",
        {
            "request": request,
            "program": program,
            "categories": categories,
            "suggestions": suggestions,
        },
    )


@router.post("/programs/{program_id}/edit")
async def program_update(
    request: Request,
    program_id: int,
    name: str = Form(...),
    category: str = Form(...),
    project_types: str = Form(""),
    selection_criteria: str = Form(""),
    permanent: str = Form(""),
    start_submission_date: str = Form(""),
    end_submission_date: str = Form(""),
    pdp_axes: str = Form(""),
    comments: str = Form(""),
    min_amount_eur: Optional[int] = Form(None),
    max_amount_eur: Optional[int] = Form(None),
    cofinancing_pct: Optional[int] = Form(None),
    application_type: str = Form(""),
    next_deadline: str = Form(""),
    eligible_structures: str = Form(""),
    eligible_themes: str = Form(""),
    source_urls: str = Form(""),
):
    """Update an existing funding program."""
    existing = await get_by_id(program_id)
    existing_urls = existing.source_urls if existing else None

    db = await get_db()
    try:
        await db.execute(
            """UPDATE funding_programs SET
                category=?, name=?, project_types=?, selection_criteria=?,
                permanent=?, start_submission_date=?, end_submission_date=?,
                pdp_axes=?, comments=?, source_urls=?,
                min_amount_eur=?, max_amount_eur=?, cofinancing_pct=?,
                eligible_structures=?, eligible_themes=?,
                application_type=?, next_deadline=?,
                last_updated_at=CURRENT_TIMESTAMP
               WHERE id=?""",
            (
                category,
                name,
                project_types,
                selection_criteria,
                bool(permanent),
                start_submission_date or None,
                end_submission_date or None,
                pdp_axes,
                comments,
                _parse_source_urls(source_urls, existing_urls),
                min_amount_eur,
                max_amount_eur,
                cofinancing_pct,
                json.dumps(_parse_comma_list(eligible_structures), ensure_ascii=False),
                json.dumps(_parse_comma_list(eligible_themes), ensure_ascii=False),
                application_type or None,
                next_deadline or None,
                program_id,
            ),
        )
        await db.commit()
    finally:
        await db.close()

    return RedirectResponse("/admin/programs", status_code=303)


@router.post("/programs/{program_id}/delete")
async def program_delete(request: Request, program_id: int):
    """Delete a funding program."""
    db = await get_db()
    try:
        await db.execute(
            "DELETE FROM funding_programs WHERE id = ?", (program_id,)
        )
        await db.commit()
    finally:
        await db.close()

    return RedirectResponse("/admin/programs", status_code=303)


# --- Backup / Export ---


@router.get("/export/json")
async def export_json():
    """Export all programs as JSON."""
    programs = await get_all()
    data = [p.model_dump(mode="json") for p in programs]
    content = json.dumps(data, ensure_ascii=False, indent=2)
    today = date.today().isoformat()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="findmyfundings-{today}.json"'
        },
    )


@router.get("/export/excel")
async def export_excel():
    """Export all programs as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    programs = await get_all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programmes"

    headers = [
        "Catégorie", "Nom", "Types de projets", "Critères de sélection",
        "Permanent", "Début soumission", "Fin soumission",
        "Axes finançables", "Commentaires",
        "Montant min (EUR)", "Montant max (EUR)", "Co-financement (%)",
        "Structures éligibles", "Thèmes éligibles",
        "Type de candidature", "Prochaine échéance", "URLs sources",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, p in enumerate(programs, 2):
        ws.cell(row=row, column=1, value=p.category)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=p.project_types)
        ws.cell(row=row, column=4, value=p.selection_criteria)
        ws.cell(row=row, column=5, value="Oui" if p.permanent else "Non")
        ws.cell(row=row, column=6, value=str(p.start_submission_date) if p.start_submission_date else "")
        ws.cell(row=row, column=7, value=str(p.end_submission_date) if p.end_submission_date else "")
        ws.cell(row=row, column=8, value=p.pdp_axes)
        ws.cell(row=row, column=9, value=p.comments)
        ws.cell(row=row, column=10, value=p.min_amount_eur)
        ws.cell(row=row, column=11, value=p.max_amount_eur)
        ws.cell(row=row, column=12, value=p.cofinancing_pct)
        ws.cell(row=row, column=13, value=", ".join(p.eligible_structures))
        ws.cell(row=row, column=14, value=", ".join(p.eligible_themes))
        ws.cell(row=row, column=15, value=p.application_type or "")
        ws.cell(row=row, column=16, value=str(p.next_deadline) if p.next_deadline else "")
        ws.cell(row=row, column=17, value=", ".join(link.url for link in p.source_urls))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="findmyfundings-{today}.xlsx"'
        },
    )


@router.get("/export/db")
async def export_db():
    """Download a copy of the SQLite database."""
    db_path = settings.db_path
    if not db_path.exists():
        return RedirectResponse("/admin/programs", status_code=303)

    buf = io.BytesIO()
    with open(db_path, "rb") as f:
        shutil.copyfileobj(f, buf)
    buf.seek(0)
    today = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/x-sqlite3",
        headers={
            "Content-Disposition": f'attachment; filename="findmyfundings-{today}.db"'
        },
    )
