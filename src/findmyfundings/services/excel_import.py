"""Import funding programs from the Excel cartography file.

Handles merged cells in column B (category) and across program rows.
Extracts hyperlinks from column I.
"""

import json
import re

import openpyxl
from openpyxl.cell.cell import MergedCell

from findmyfundings.database import get_db

_AMOUNT_RE = re.compile(
    r'(\d[\d\s.,]*)\s*(k€|K€|k ?euros?|K ?euros?|M€|M ?euros?|€|euros?)',
    re.IGNORECASE,
)


def _parse_euro_amount(number_str: str, unit: str) -> int | None:
    """Convert a matched amount string to euros."""
    try:
        cleaned = number_str.strip().replace(" ", "").replace(",", ".")
        value = float(cleaned)
        unit_lower = unit.lower().replace(" ", "")
        if unit_lower.startswith("m"):
            return int(value * 1_000_000)
        if unit_lower.startswith("k"):
            return int(value * 1_000)
        return int(value)
    except (ValueError, OverflowError):
        return None


def _extract_amounts(text: str) -> tuple[int | None, int | None]:
    """Extract min and max euro amounts from free text."""
    matches = _AMOUNT_RE.findall(text)
    if not matches:
        return None, None
    amounts = []
    for num_str, unit in matches:
        val = _parse_euro_amount(num_str, unit)
        if val and val > 0:
            amounts.append(val)
    if not amounts:
        return None, None
    return min(amounts), max(amounts)


def _get_merged_value(ws, cell):
    """Get the value of a cell, resolving merged cell references."""
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def _collect_hyperlinks(ws, row_num, col=9):
    """Collect hyperlink from column I of a given row."""
    cell = ws.cell(row_num, col)
    links = []
    if cell.hyperlink and cell.hyperlink.target:
        label = str(cell.value or "").strip()
        if label.startswith("http"):
            label = ""
        links.append({"url": cell.hyperlink.target, "label": label})
    elif cell.value and str(cell.value).startswith("http"):
        links.append({"url": str(cell.value).strip(), "label": ""})
    return links


def parse_excel(file_path: str) -> list[dict]:
    """Parse the Excel file and return a list of funding program dicts."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    programs = []
    current_category = ""

    # Find which rows are the "start" of merged ranges in column C (program name)
    # to handle multi-row programs
    name_merged_ranges = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 3:  # Column C
            name_merged_ranges[mr.min_row] = mr.max_row

    row = 4  # Skip header rows (1-3)
    while row <= ws.max_row:
        # Get category from column B (handles merged cells)
        cat_cell = ws.cell(row, 2)
        cat_value = _get_merged_value(ws, cat_cell)
        if cat_value:
            current_category = str(cat_value).strip()

        # Get program name from column C
        name_cell = ws.cell(row, 3)
        name_value = _get_merged_value(ws, name_cell)

        if not name_value:
            # Check if this row has any data at all (some rows are info/header rows)
            has_data = any(
                ws.cell(row, col).value for col in range(4, 10)
            )
            if not has_data:
                row += 1
                continue
            # Row with data but no name - might be a category info row
            # Check column I for links
            links = _collect_hyperlinks(ws, row)
            if links and current_category:
                # Store as a reference/info entry for the category
                programs.append({
                    "category": current_category,
                    "name": f"[Info] {current_category}",
                    "project_types": str(ws.cell(row, 4).value or ""),
                    "selection_criteria": str(ws.cell(row, 5).value or ""),
                    "submission_dates": str(ws.cell(row, 6).value or ""),
                    "pdp_axes": str(ws.cell(row, 7).value or ""),
                    "comments": str(ws.cell(row, 8).value or ""),
                    "source_urls": links,
                })
            row += 1
            continue

        name = str(name_value).strip()

        # Determine how many rows this program spans
        end_row = name_merged_ranges.get(row, row)

        # Collect data across all rows of this program
        project_types_parts = []
        criteria_parts = []
        dates_parts = []
        axes_parts = []
        comments_parts = []
        all_links = []

        for r in range(row, end_row + 1):
            for col, parts in [
                (4, project_types_parts),
                (5, criteria_parts),
                (6, dates_parts),
                (7, axes_parts),
                (8, comments_parts),
            ]:
                val = ws.cell(r, col).value
                if val is None:
                    # Check if it's a merged cell
                    val = _get_merged_value(ws, ws.cell(r, col))
                if val and str(val).strip():
                    text = str(val).strip()
                    if text not in parts:
                        parts.append(text)

            all_links.extend(_collect_hyperlinks(ws, r))

        project_types_text = "\n".join(project_types_parts)
        criteria_text = "\n".join(criteria_parts)
        amount_text = project_types_text + " " + criteria_text
        min_amount, max_amount = _extract_amounts(amount_text)

        programs.append({
            "category": current_category,
            "name": name,
            "project_types": project_types_text,
            "selection_criteria": criteria_text,
            "submission_dates": "\n".join(dates_parts),
            "pdp_axes": "\n".join(axes_parts),
            "comments": "\n".join(comments_parts),
            "source_urls": all_links,
            "min_amount_eur": min_amount,
            "max_amount_eur": max_amount,
        })

        row = end_row + 1

    return programs


async def import_to_db(programs: list[dict]):
    """Insert parsed programs into the database."""
    db = await get_db()
    try:
        # Clear existing data
        await db.execute("DELETE FROM monitored_sources")
        await db.execute("DELETE FROM funding_programs")
        await db.execute("DELETE FROM funding_fts")

        for prog in programs:
            source_urls_json = json.dumps(prog["source_urls"], ensure_ascii=False)
            cursor = await db.execute(
                """INSERT INTO funding_programs
                   (category, name, project_types, selection_criteria,
                    submission_dates, pdp_axes, comments, source_urls,
                    min_amount_eur, max_amount_eur)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    prog["category"],
                    prog["name"],
                    prog["project_types"],
                    prog["selection_criteria"],
                    prog["submission_dates"],
                    prog["pdp_axes"],
                    prog["comments"],
                    source_urls_json,
                    prog.get("min_amount_eur"),
                    prog.get("max_amount_eur"),
                ),
            )
            program_id = cursor.lastrowid

            # Register source URLs for monitoring
            for link in prog["source_urls"]:
                if link["url"].startswith("http"):
                    try:
                        await db.execute(
                            """INSERT OR IGNORE INTO monitored_sources
                               (url, label, funding_program_id)
                               VALUES (?, ?, ?)""",
                            (link["url"], link["label"], program_id),
                        )
                    except Exception:
                        pass  # Skip duplicate URLs

        await db.commit()
        return len(programs)
    finally:
        await db.close()
